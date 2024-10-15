import openpyxl

state2Code ={'alabama': 'al', 'alaska': 'ak', 'arizona': 'az', 'arkansas': 'ar', 'california': 'ca', 'colorado': 'co', 'connecticut': 'ct', 'delaware': 'de', 'florida': 'fl', 'georgia': 'ga', 'hawaii': 
'hi', 'idaho': 'id', 'illinois': 'il', 'indiana': 'in', 'iowa': 'ia', 'kansas': 'ks', 'kentucky': 'ky', 'louisiana': 'la', 'maine': 'me', 'maryland': 'md', 'massachusetts': 'ma', 'michigan': 'mi', 'minnesota': 'mn', 'mississippi': 'ms', 'missouri': 'mo', 'montana': 'mt', 'nebraska': 'ne', 'nevada': 'nv', 'newhampshire': 'nh', 'newjersey': 'nj', 'newmexico': 'nm', 'newyork': 'ny', 'northcarolina': 'nc', 'northdakota': 'nd', 'ohio': 'oh', 'oklahoma': 'ok', 'oregon': 'or', 
'pennsylvania': 'pa', 'rhodeisland': 'ri', 'southcarolina': 'sc', 'southdakota': 'sd', 'tennessee': 'tn', 'texas': 'tx', 'utah': 'ut', 'vermont': 'vt', 'virginia': 'va', 'washington': 'wa', 
'westvirginia': 'wv', 'wisconsin': 'wi', 'wyoming': 'wy', 'districtofcolumbia': 'dc', 'puertorico': 'pr', 'virginislands,u.s.': 'vi'}

code2State ={'al': 'alabama', 'ak': 'alaska', 'az': 'arizona', 'ar': 'arkansas', 'ca': 'california', 'co': 
'colorado', 'ct': 'connecticut', 'de': 'delaware', 'fl': 'florida', 'ga': 'georgia', 'hi': 'hawaii', 'id': 'idaho', 'il': 'illinois', 'in': 'indiana', 'ia': 'iowa', 'ks': 'kansas', 'ky': 'kentucky', 'la': 'louisiana', 'me': 'maine', 'md': 'maryland', 'ma': 'massachusetts', 'mi': 'michigan', 'mn': 'minnesota', 'ms': 'mississippi', 'mo': 'missouri', 'mt': 'montana', 'ne': 'nebraska', 'nv': 'nevada', 'nh': 'newhampshire', 'nj': 'newjersey', 'nm': 'new-mexico', 'ny': 'new-york', 'nc': 'north-carolina', 'nd': 'north-dakota', 'oh': 'ohio', 'ok': 'oklahoma', 'or': 'oregon', 
'pa': 'pennsylvania', 'ri': 'rhode-island', 'sc': 'south-carolina', 'sd': 'south-dakota', 'tn': 'tennessee', 'tx': 'texas', 'ut': 'utah', 'vt': 'vermont', 'va': 'virginia', 'wa': 'washington', 
'wv': 'west-virginia', 'wi': 'wisconsin', 'wy': 'wyoming', 'dc': 'washington-dc', 'pr': 'puerto-rico', 'vi': 'virginislands,u.s.'}

insuranceReasonTokens ={
    'high annual maximum': "with-a-high-annual-maximum",
    'low coinsurance':"for-major-dental-work",
    'dental treatment soon':'with-no-waiting-period',
    'having dental implants covered':'for-dental-implants',
    'braces or clear aligners covered':'for-braces',
    'having teeth whitening covered':'for-teeth-whitening'
}

def getEmailLine(emailBody, textSearch):
    for i in range(0, len(emailBody)):
        if textSearch in emailBody[i]:
            return i
        
    return -2

def parseEmail(emailBody):
    formDict = {
        'stateCode':None,
        'county': None,
        'city': None,
        'insuranceOrMedicarePlanBool': None,
        'insuranceOrMedicarePlan': None,
        'recommendInsuranceBool': None,
        'importantDentalInsurance': None,
        'proceduresNeeded': None,
        'howSoon': None,
        'days': None,
        'timeOfDay': None,
        'whatToDo': None,
        'email': None
    }
    # print("I received the following data")
    # print(emailBody)

    #find state
    
    inputRaw = emailBody[getEmailLine(emailBody, "Which state do you live in?")+1].strip().lower().replace(' ', '')
    if not inputRaw.isalpha(): return "Error in state name: "+inputRaw
    for key in state2Code:
        if inputRaw in key:
            formDict['stateCode'] = state2Code[key]
            break

    inputRaw = emailBody[getEmailLine(emailBody, "Which county do you live in?")+1].strip().lower()
    formDict['county'] = inputRaw

    inputRaw = emailBody[getEmailLine(emailBody, "Which city do you live in?")+1].strip().lower()
    if not inputRaw.isalpha(): return "Error in city name: "+inputRaw
    formDict['city'] = inputRaw

    inputRaw = emailBody[getEmailLine(emailBody, "Do you have dental insurance or a Medicare Advantage plan?")+1].strip().lower()
    if not inputRaw.isalpha(): return "Error in dental insurance or medicare answer: "+inputRaw
    if inputRaw == "yes":
        formDict['insuranceOrMedicarePlanBool'] = True
    else:
        formDict['insuranceOrMedicarePlanBool'] = False

    #yes answers
    try:
        inputRaw = emailBody[getEmailLine(emailBody, "Please select your dental insurance or Medicare Advantage plan")+1].strip().lower()
        # if not inputRaw.isalpha(): return "Error in Dental insurance Medicare name: "+inputRaw
        if "Please specify" in inputRaw:
            inputRaw = inputRaw[len("Please specify"):].strip()
        formDict['insuranceOrMedicarePlan'] = inputRaw
    except:
        pass

    #no answers
    try:
        inputRaw = emailBody[getEmailLine(emailBody, "Would you like us to recommend the best dental insurance for your needs?")+1].strip().lower()
        # if not inputRaw.isalpha(): return "Error in recommend the best dental insurance for your needs: "+inputRaw
        if len(inputRaw):
            if inputRaw == "yes":
                formDict['recommendInsuranceBool'] = True
            else:
                formDict['recommendInsuranceBool'] = False
    except:
        pass

    try:
        inputRaw = emailBody[getEmailLine(emailBody, "What is most important to you about having dental insurance?")+1].strip().lower()
        if len(inputRaw):
            # if not inputRaw.isalpha(): return "Error in importantDentalInsurance: "+inputRaw
            formDict['importantDentalInsurance'] = inputRaw
    except:
        pass

    try:
        inputRaw = emailBody[getEmailLine(emailBody, "Choose which dental procedure you think you will need soon")+1].strip().lower()
        # if not inputRaw.isalpha(): return "Error in proceduresNeeded: "+inputRaw
        formDict['proceduresNeeded'] = inputRaw.split(',')
        for i in range(len(formDict['proceduresNeeded'])):
            formDict['proceduresNeeded'][i] = formDict['proceduresNeeded'][i].strip()
    except:
        pass

    inputRaw = emailBody[getEmailLine(emailBody, "How soon do you want to see a dentist?")+1].strip().lower()
    # if not inputRaw.isalpha(): return "Error in howSoon: "+inputRaw
    formDict['howSoon'] = inputRaw

    inputRaw = emailBody[getEmailLine(emailBody, "Which day of the week works best for a dental appointment?")+1].strip().lower()
    formDict['days'] = inputRaw

    inputRaw = emailBody[getEmailLine(emailBody, "What time of day do you prefer seeing a dentist?")+1].strip().lower()
    formDict['timeOfDay'] = inputRaw

    inputRaw = emailBody[getEmailLine(emailBody, "Select how you would like us to help you find a dentist.")+1].strip().lower()
    # if not inputRaw.isalpha(): return "Error in whatToDo: "+inputRaw
    formDict['whatToDo'] = inputRaw

    inputRaw = emailBody[getEmailLine(emailBody, "In order to receive your personalized results, please enter your email.")+1].strip().lower()
    formDict['email'] = inputRaw

    return formDict



    print(stateRaw)

def checkTokensInLink(link, tokens):
    for token in tokens:
        if token not in link:
            return False
    
    return True

def getSection2Links(stateCode, county, city):
    print("In getSection2Links")
    print("received", stateCode, county, city)
    thisDict = {
        'cityMatch': False,
        'countyMatch': False,
        'cityLink': None,
        'countyLink':None
    }

    #start countyLookup
    wb_obj = openpyxl.load_workbook("./devotedCounties.xlsx")
    sheet_obj = wb_obj.active
    totalRows = sheet_obj.max_row
    for i in range(1, totalRows+1):
        if str(sheet_obj.cell(row = i, column = 1).value) == stateCode and str(sheet_obj.cell(row = i, column = 2).value).lower() in county:
            thisDict['countyMatch']= True


    #cityMatch
    wb_obj = openpyxl.load_workbook("./devotedCities.xlsx")
    sheet_obj = wb_obj.active
    totalColumns = sheet_obj.max_column
    totalRows = sheet_obj.max_row
    print("Let's check the state codes in city list")
    for i in range(1, totalColumns+1):
        print(str(sheet_obj.cell(row = 1, column = i).value))
        if str(sheet_obj.cell(row = 1, column = i).value).lower() == stateCode:
            print("Received a stateCode match. Checking cities now")
            for j in range(1, totalRows+1):
                print(str(sheet_obj.cell(row = j, column = i).value))
                if str(sheet_obj.cell(row = j, column = i).value).lower().replace(".", "").replace(" ", "") == city:
                    print("I got city match")
                    thisDict['cityMatch'] = True
                    break
        if thisDict['cityMatch']:
            break

    if thisDict['cityMatch']:
        #begin match in sitemapCut
        f = open("./sitemapFull.txt", "r")
        links = f.readlines()
        f.close()

        tokens = ["best-dentists-in"]
        tokens.append(stateCode)
        cityToken = city.lower().replace(".", "").replace(" ", "-")
        tokens.append(cityToken)

        for link in links:
            if checkTokensInLink(link, tokens):
                thisDict['cityLink'] = link
                
            if thisDict['cityLink'] is not None:
                break

    return thisDict


def findInsuranceLink(tokens):
    f = open("./sitemapFull.txt", "r")
    completeList = f.readlines()
    f.close()

    for link in completeList:
        if tokens[0] in link and tokens[1] in link:
            return link
    return None


def getSection3Links(stateCode, importantReasons):
    links = []


    for key in insuranceReasonTokens:
        if key in importantReasons:
            linkFromList = findInsuranceLink([insuranceReasonTokens[key], code2State[stateCode]])
            if linkFromList:
                links.append(linkFromList)


    return links

def buildEmail(formDict):

    emailText = ""
    section1Text = """
Howdy! We're glad you found us and we'll do our best to help!<br><br>
"""

    emailText += section1Text
    
    #section 2
    if "denscore to email me a list of the best dentists" in formDict['whatToDo']:
        section2Text = "Based on the information you provided, we have included a list of the best dentists near you below.<br>"
        linksDict = getSection2Links(formDict['stateCode'], formDict['county'], formDict['city'])
        if linksDict['cityMatch']:
            section2Text+=("<a href=\""+linksDict['cityLink']+"\">"+linksDict['cityLink']+"</a><br>")

        if linksDict['countyMatch']:
            section2Text+=("<a href=\""+linksDict['countyLink']+"\">"+linksDict['countyLink']+"</a><br>")

        if not linksDict['cityMatch'] and not linksDict['countyMatch']:
            section2Text = """
Based on the information you provided, we don't have a list of the best dentists near you at this time. However, we've included a list of things you should know before choosing a dentist below. 
<a href="https://www.denscore.com/how-do-i-find-the-best-dentist-near-me/">https://www.denscore.com/how-do-i-find-the-best-dentist-near-me/</a><br>
"""
    else:
        section2Text = "You indicated that you'd like us to call the best dental office nearby and have them email you information about scheduling a dental appointment. Within 24 hours, you'll receive an email from us with the name of the dental office that will be reaching out to you along with other pertinent information so stay tuned!<br>"

    emailText+=section2Text
    emailText+="<br>"
    #section 3 
    if formDict['recommendInsuranceBool'] and len(formDict['importantDentalInsurance']):
        section3Text = "Because you asked us to recommend a dental insurance plan, we've included a list of the best dental insurance plans below based on your needs.<br>"
        links = getSection3Links(formDict['stateCode'], formDict['importantDentalInsurance'])

        for link in links:
            section3Text+=("<a href=\""+link+"\">"+link+"</a><br>")

        emailText+= section3Text
        emailText+="<br>"

    #section 4
    if len(formDict['proceduresNeeded']):
        section4Text = """
 You mentioned that you may need some dental work, so we've included one or more links below which provide important information about the procedures you listed in your form. Be sure to have a look because these tips could save you money and allow you to make a more informed decision about your dental care!<br>
"""

        if "filling" in formDict['proceduresNeeded']:
            toAppend = """
<a href="https://www.denscore.com/white-or-silver-fillings-which-is-best-for-me/">https://www.denscore.com/white-or-silver-fillings-which-is-best-for-me/</a><br>
"""
            section4Text+=(toAppend)


        if "crown" in formDict['proceduresNeeded']:
            toAppend = """
<a href="https://www.denscore.com/choosing-the-right-crown-for-your-smile/">https://www.denscore.com/choosing-the-right-crown-for-your-smile/</a><br>
<a href="https://www.denscore.com/veneers-lumineers-and-crowns-a-comprehensive-guide/">https://www.denscore.com/veneers-lumineers-and-crowns-a-comprehensive-guide/</a><br>
<a href="https://www.denscore.com/why-do-my-gums-hurt-after-getting-a-crown/">https://www.denscore.com/why-do-my-gums-hurt-after-getting-a-crown/</a><br>
"""
            section4Text+=(toAppend)

        if "bridge" in formDict['proceduresNeeded']:
            toAppend = """
<a href="https://www.denscore.com/dental-implant-vs-bridge-for-a-missing-tooth-a-comparison/">https://www.denscore.com/dental-implant-vs-bridge-for-a-missing-tooth-a-comparison</a><br>
<a href="https://www.denscore.com/advantages-of-maryland-bridges-for-tooth-replacement/">https://www.denscore.com/advantages-of-maryland-bridges-for-tooth-replacement/</a><br>
"""
            section4Text+=(toAppend)

        if "extraction" in formDict['proceduresNeeded']:
            toAppend = """
<a href="https://www.denscore.com/navigating-dental-decisions-root-canal-vs-tooth-extraction/">https://www.denscore.com/navigating-dental-decisions-root-canal-vs-tooth-extraction/</a><br>
<a href="https://www.denscore.com/decoding-bone-grafts-navigating-the-essentials/">https://www.denscore.com/decoding-bone-grafts-navigating-the-essentials/</a><br>
"""
            section4Text+=(toAppend)

        if "denture" in formDict['proceduresNeeded']:
            toAppend = """
<a href="https://www.denscore.com/smile-confidently-a-guide-to-different-types-of-dentures/">https://www.denscore.com/smile-confidently-a-guide-to-different-types-of-dentures/</a><br>
<a href="https://www.denscore.com/denture-care-a-comprehensive-guide/">https://www.denscore.com/denture-care-a-comprehensive-guide/</a><br>
<a href="https://www.denscore.com/where-do-i-find-affordable-dentures-and-implants-near-me/">https://www.denscore.com/where-do-i-find-affordable-dentures-and-implants-near-me/</a><br>
"""
            section4Text+=(toAppend)


        if "dental implant" in formDict['proceduresNeeded']:
            toAppend = """
<a href="https://www.denscore.com/how-much-are-dental-implants/">https://www.denscore.com/how-much-are-dental-implants/</a><br>
<a href="https://www.denscore.com/how-long-do-dental-implants-last/">https://www.denscore.com/how-long-do-dental-implants-last/</a><br>
<a href="https://www.denscore.com/screw-retained-vs-cement-retained-dental-implant-crowns/">https://www.denscore.com/screw-retained-vs-cement-retained-dental-implant-crowns/</a><br>
"""
            section4Text+=(toAppend)

        if "teeth straightening" in formDict['proceduresNeeded']:
            toAppend = """
<a href="https://www.denscore.com/5-essential-steps-before-straightening-your-teeth/">https://www.denscore.com/5-essential-steps-before-straightening-your-teeth/</a><br>
<a href="https://www.denscore.com/in-person-orthodontic-care-vs-mail-order-aligners/">https://www.denscore.com/in-person-orthodontic-care-vs-mail-order-aligners/</a><br>
"""
            section4Text+=(toAppend)


        if "teeth whitening" in formDict['proceduresNeeded']:
            toAppend = """
<a href="https://www.denscore.com/teeth-whitening-a-guide-to-brighter-smiles/">https://www.denscore.com/teeth-whitening-a-guide-to-brighter-smiles/</a><br>
<a href="https://www.denscore.com/unveiling-the-top-3-over-the-counter-teeth-whitening-kits/">https://www.denscore.com/unveiling-the-top-3-over-the-counter-teeth-whitening-kits/</a><br>
"""
            section4Text+=(toAppend)

        if "deep cleaning or gum surgery" in formDict['proceduresNeeded']:
            toAppend = """
<a href="https://www.denscore.com/gum-disease-treatments-early-intervention-to-advanced-solutions/">https://www.denscore.com/gum-disease-treatments-early-intervention-to-advanced-solutions/</a><br>
<a href="https://www.denscore.com/guide-to-the-management-of-periodontal-disease-by-denscore/">https://www.denscore.com/guide-to-the-management-of-periodontal-disease-by-denscore/</a><br>
<a href="https://www.denscore.com/gum-abscess-vs-periodontal-abscess-understanding-the-difference/">https://www.denscore.com/gum-abscess-vs-periodontal-abscess-understanding-the-difference/</a><br>
"""
            section4Text+=(toAppend)

        if "dental exam/cleaning" in formDict['proceduresNeeded']:
            toAppend = """
<a href="https://www.denscore.com/causes-of-plaque-and-how-to-prevent-plaque-build-up/">https://www.denscore.com/causes-of-plaque-and-how-to-prevent-plaque-build-up/</a><br>
"""
            section4Text+=(toAppend)

        emailText+= section4Text
        emailText+= "<br>"


    lastSection = """
    We hope this information was helpful. If you asked us to call the best dentist and you don't hear from the office within 2 business days, or if you have specific requests (ex. "Monday afternoons work best for a dental appointment" etc.), please <a href="mailto:notifications@denscore.com">email us here</a>. 
    <br><br>Very Respectfully,
    """

    emailText+= lastSection
    emailText+= "<br>"

    return emailText